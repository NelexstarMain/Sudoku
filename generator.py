import random
from typing import List, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

class Sudoku:
    CLUES = {1: 60, 2: 55, 3: 45, 4: 30, 5: 20}

    def __init__(self, lvl: int = 1) -> None:
        self.lvl = lvl
        self.board: List[List[int]] = [[0] * 9 for _ in range(9)]
        self.solution: List[List[int]] = []  
        self.wb = Workbook()  

    def _valid(self, num: int, pos: Tuple[int, int]) -> bool:
        r, c = pos
        if any(self.board[r][j] == num for j in range(9) if j != c):
            return False
        if any(self.board[i][c] == num for i in range(9) if i != r):
            return False
        br, bc = (r // 3) * 3, (c // 3) * 3
        for i in range(br, br + 3):
            for j in range(bc, bc + 3):
                if (i, j) != pos and self.board[i][j] == num:
                    return False
        return True

    def _find_empty(self) -> Optional[Tuple[int, int]]:
        for i in range(9):
            for j in range(9):
                if self.board[i][j] == 0:
                    return (i, j)
        return None

    def _solve(self) -> bool:
        empty = self._find_empty()
        if not empty:
            return True
        r, c = empty
        nums = list(range(1, 10))
        random.shuffle(nums)
        for num in nums:
            if self._valid(num, (r, c)):
                self.board[r][c] = num
                if self._solve():
                    return True
                self.board[r][c] = 0
        return False

    def generate(self) -> None:
        """
        Generuje pełną planszę sudoku, zapisuje jej kopię jako rozwiązanie,
        a następnie usuwa komórki według wybranego poziomu trudności, tworząc łamigłówkę.
        """
        if not self._solve():
            raise RuntimeError("Nie udało się wygenerować planszy sudoku.")
        self.solution = [row[:] for row in self.board]
        cells_to_remove = 81 - self.CLUES.get(self.lvl, 60)
        while cells_to_remove > 0:
            i, j = random.randint(0, 8), random.randint(0, 8)
            if self.board[i][j] != 0:
                self.board[i][j] = 0
                cells_to_remove -= 1

    def _format_sheet(self, ws, board: List[List[int]]) -> None:
        """
        Formatuje arkusz Excela tak, aby plansza wyglądała jak klasyczne sudoku:
         - Ustawia wartości, wyśrodkowanie i pogrubioną czcionkę.
         - Rysuje cienkie i grube granice oddzielające bloki 3x3.
         - Ustawia odpowiednią szerokość kolumn i wysokość wierszy.
        """
        thin = Side(style='thin', color="000000")
        thick = Side(style='thick', color="000000")
        for i in range(9):
            for j in range(9):
                cell = ws.cell(row=i+1, column=j+1)
                val = board[i][j]
                cell.value = val if val != 0 else ""
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=14, bold=True)
                left_border   = thick if j % 3 == 0 else thin
                right_border  = thick if (j + 1) % 3 == 0 else thin
                top_border    = thick if i % 3 == 0 else thin
                bottom_border = thick if (i + 1) % 3 == 0 else thin
                cell.border = Border(left=left_border,
                                     right=right_border,
                                     top=top_border,
                                     bottom=bottom_border)
                ws.column_dimensions[cell.column_letter].width = 4
            ws.row_dimensions[i+1].height = 30

    def excel(self, filename: str = "sudoku.xlsx") -> None:
        """
        Zapisuje łamigłówkę (Puzzle) oraz pełne rozwiązanie (Solution) do pliku Excel.
        """
        puzzle_sheet = self.wb.active
        puzzle_sheet.title = "Puzzle"
        self._format_sheet(puzzle_sheet, self.board)

        solution_sheet = self.wb.create_sheet(title="Solution")
        self._format_sheet(solution_sheet, self.solution)

        self.wb.save(filename)
        print(f"Plansza sudoku zapisana do pliku '{filename}' z arkuszami: 'Puzzle' i 'Solution'.")

    def __repr__(self) -> str:
        lines = []
        for i, row in enumerate(self.board):
            if i % 3 == 0 and i != 0:
                lines.append("-" * 21)
            line = ""
            for j, val in enumerate(row):
                if j % 3 == 0 and j != 0:
                    line += "| "
                line += f"{val if val else '.'} "
            lines.append(line)
        return "\n".join(lines)

if __name__ == "__main__":
    sudoku = Sudoku(lvl=2)
    sudoku.generate()

    print("Puzzle:")
    print(sudoku)

    print("\nSolution:")
    solution_text = "\n".join(
        " ".join(str(cell) if cell != 0 else '.' for cell in row)
        for row in sudoku.solution
    )
    print(solution_text)

    sudoku.excel()
